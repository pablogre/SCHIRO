#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cta_cte.py - MÓDULO DE CUENTA CORRIENTE
═══════════════════════════════════════════════════════════════════════════════
Maneja toda la lógica de ventas fiadas y pagos en cuenta corriente.
═══════════════════════════════════════════════════════════════════════════════
"""

from flask import Blueprint, jsonify, request, session, render_template
from sqlalchemy import text, func
from datetime import datetime
from decimal import Decimal

# Importar auditoría de stock
try:
    from stock_audit import registrar_movimiento_stock
except ImportError:
    # Si no existe el módulo, crear función dummy
    def registrar_movimiento_stock(*args, **kwargs):
        pass

# Blueprint para las rutas de CTA.CTE
cta_cte_bp = Blueprint('cta_cte', __name__)


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════════

def ejecutar_query(db, query, params=None, commit=False):
    """
    Ejecuta una query SQL de forma segura
    
    Args:
        db: Instancia de SQLAlchemy
        query: Query SQL (puede usar :param para parámetros)
        params: Diccionario de parámetros
        commit: Si True, hace commit
    
    Returns:
        ResultProxy o lastrowid según el caso
    """
    try:
        if params:
            result = db.session.execute(text(query), params)
        else:
            result = db.session.execute(text(query))
        
        if commit:
            db.session.commit()
            return result.lastrowid if hasattr(result, 'lastrowid') else True
        
        return result
    except Exception as e:
        db.session.rollback()
        raise e


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES PRINCIPALES - GUARDAR VENTA FIADA
# ═══════════════════════════════════════════════════════════════════════════════

def guardar_venta_fiada(db, cliente_id, productos, usuario_id, observaciones=None):
    """
    Guarda una venta fiada en cuenta corriente
    - NO emite factura
    - SÍ descuenta stock (incluyendo combos)
    - Guarda como pendiente de pago
    """
    try:
        # Calcular total
        monto_total = sum(Decimal(str(p['subtotal'])) for p in productos)
        
        # 1. Crear el movimiento en CTA.CTE
        query_movimiento = """
            INSERT INTO cta_cte_movimiento 
            (cliente_id, tipo, estado, monto_total, usuario_id, observaciones, fecha)
            VALUES 
            (:cliente_id, 'venta_fiada', 'pendiente', :monto_total, :usuario_id, :observaciones, NOW())
        """
        
        result = ejecutar_query(db, query_movimiento, {
            'cliente_id': cliente_id,
            'monto_total': float(monto_total),
            'usuario_id': usuario_id,
            'observaciones': observaciones
        }, commit=True)
        
        movimiento_id = result
        
        # 2. Insertar los detalles de productos
        query_detalle = """
            INSERT INTO cta_cte_detalle 
            (movimiento_id, producto_id, descripcion, cantidad, precio_unitario, subtotal, porcentaje_iva, importe_iva)
            VALUES 
            (:movimiento_id, :producto_id, :descripcion, :cantidad, :precio_unitario, :subtotal, :porcentaje_iva, :importe_iva)
        """
        
        for producto in productos:
            ejecutar_query(db, query_detalle, {
                'movimiento_id': movimiento_id,
                'producto_id': producto['producto_id'],
                'descripcion': producto['descripcion'],
                'cantidad': float(producto['cantidad']),
                'precio_unitario': float(producto['precio_unitario']),
                'subtotal': float(producto['subtotal']),
                'porcentaje_iva': float(producto.get('porcentaje_iva', 21.00)),
                'importe_iva': float(producto.get('importe_iva', 0.00))
            }, commit=True)
        
        # 3. Descontar stock - CORREGIDO PARA COMBOS CON AUDITORÍA
        for producto in productos:
            producto_id = producto['producto_id']
            cantidad = float(producto['cantidad'])
            
            # Verificar si es combo
            query_check_combo = """
                SELECT id, codigo, nombre, es_combo, stock, producto_base_id, cantidad_combo,
                       producto_base_id_2, cantidad_combo_2,
                       producto_base_id_3, cantidad_combo_3
                FROM producto WHERE id = :producto_id
            """
            result_combo = ejecutar_query(db, query_check_combo, {'producto_id': producto_id})
            row = result_combo.fetchone()
            
            if row and row.es_combo:
                # Es combo - descontar de productos base
                if row.producto_base_id and row.cantidad_combo:
                    descuento = cantidad * float(row.cantidad_combo)
                    
                    # Obtener info del producto base para auditoría
                    query_base_info = "SELECT id, codigo, nombre, stock FROM producto WHERE id = :base_id"
                    result_base = ejecutar_query(db, query_base_info, {'base_id': row.producto_base_id})
                    base_info = result_base.fetchone()
                    
                    if base_info:
                        stock_anterior = float(base_info.stock)
                        stock_nuevo = stock_anterior - descuento
                        
                        query_stock_base = """
                            UPDATE producto 
                            SET stock = stock - :descuento 
                            WHERE id = :base_id
                        """
                        ejecutar_query(db, query_stock_base, {
                            'descuento': descuento,
                            'base_id': row.producto_base_id
                        }, commit=True)
                        print(f"📦 Combo: descontado {descuento} de producto base {row.producto_base_id}")
                        
                        # Auditoría
                        registrar_movimiento_stock(
                            db=db,
                            producto_id=base_info.id,
                            tipo='venta_fiada',
                            cantidad=descuento,
                            signo='-',
                            stock_anterior=stock_anterior,
                            stock_nuevo=stock_nuevo,
                            referencia_tipo='cta_cte',
                            referencia_id=movimiento_id,
                            motivo=f'Combo en CTA.CTE',
                            usuario_id=session.get('user_id'),
                            usuario_nombre=session.get('nombre', 'Sistema'),
                            codigo_producto=base_info.codigo,
                            nombre_producto=base_info.nombre
                        )
                
                if row.producto_base_id_2 and row.cantidad_combo_2:
                    descuento = cantidad * float(row.cantidad_combo_2)
                    
                    result_base = ejecutar_query(db, query_base_info, {'base_id': row.producto_base_id_2})
                    base_info = result_base.fetchone()
                    
                    if base_info:
                        stock_anterior = float(base_info.stock)
                        stock_nuevo = stock_anterior - descuento
                        
                        ejecutar_query(db, query_stock_base, {
                            'descuento': descuento,
                            'base_id': row.producto_base_id_2
                        }, commit=True)
                        print(f"📦 Combo: descontado {descuento} de producto base 2 {row.producto_base_id_2}")
                        
                        registrar_movimiento_stock(
                            db=db,
                            producto_id=base_info.id,
                            tipo='venta_fiada',
                            cantidad=descuento,
                            signo='-',
                            stock_anterior=stock_anterior,
                            stock_nuevo=stock_nuevo,
                            referencia_tipo='cta_cte',
                            referencia_id=movimiento_id,
                            motivo=f'Combo en CTA.CTE',
                            usuario_id=session.get('user_id'),
                            usuario_nombre=session.get('nombre', 'Sistema'),
                            codigo_producto=base_info.codigo,
                            nombre_producto=base_info.nombre
                        )
                
                if row.producto_base_id_3 and row.cantidad_combo_3:
                    descuento = cantidad * float(row.cantidad_combo_3)
                    
                    result_base = ejecutar_query(db, query_base_info, {'base_id': row.producto_base_id_3})
                    base_info = result_base.fetchone()
                    
                    if base_info:
                        stock_anterior = float(base_info.stock)
                        stock_nuevo = stock_anterior - descuento
                        
                        ejecutar_query(db, query_stock_base, {
                            'descuento': descuento,
                            'base_id': row.producto_base_id_3
                        }, commit=True)
                        print(f"📦 Combo: descontado {descuento} de producto base 3 {row.producto_base_id_3}")
                        
                        registrar_movimiento_stock(
                            db=db,
                            producto_id=base_info.id,
                            tipo='venta_fiada',
                            cantidad=descuento,
                            signo='-',
                            stock_anterior=stock_anterior,
                            stock_nuevo=stock_nuevo,
                            referencia_tipo='cta_cte',
                            referencia_id=movimiento_id,
                            motivo=f'Combo en CTA.CTE',
                            usuario_id=session.get('user_id'),
                            usuario_nombre=session.get('nombre', 'Sistema'),
                            codigo_producto=base_info.codigo,
                            nombre_producto=base_info.nombre
                        )
            else:
                # Producto normal - descontar directo
                stock_anterior = float(row.stock) if row else 0
                stock_nuevo = stock_anterior - cantidad
                
                query_stock = """
                    UPDATE producto 
                    SET stock = stock - :cantidad 
                    WHERE id = :producto_id
                """
                ejecutar_query(db, query_stock, {
                    'cantidad': cantidad,
                    'producto_id': producto_id
                }, commit=True)
                print(f"📦 Stock descontado: {cantidad} de producto {producto_id}")
                
                # Auditoría
                if row:
                    registrar_movimiento_stock(
                        db=db,
                        producto_id=producto_id,
                        tipo='venta_fiada',
                        cantidad=cantidad,
                        signo='-',
                        stock_anterior=stock_anterior,
                        stock_nuevo=stock_nuevo,
                        referencia_tipo='cta_cte',
                        referencia_id=movimiento_id,
                        usuario_id=session.get('user_id'),
                        usuario_nombre=session.get('nombre', 'Sistema'),
                        codigo_producto=row.codigo,
                        nombre_producto=row.nombre
                    )
        
        return {
            'success': True,
            'movimiento_id': movimiento_id,
            'mensaje': f'Venta fiada registrada correctamente. Total: ${monto_total:.2f}'
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en guardar_venta_fiada: {str(e)}")
        return {
            'success': False,
            'mensaje': f'Error al guardar venta fiada: {str(e)}'
        }
        

# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES PRINCIPALES - OBTENER PRODUCTOS PENDIENTES
# ═══════════════════════════════════════════════════════════════════════════════

def obtener_productos_pendientes(db, cliente_id):
    """
    Obtiene todos los productos pendientes de pago de un cliente
    
    Args:
        db: Instancia de SQLAlchemy
        cliente_id: ID del cliente
    
    Returns:
        Lista de dict con productos pendientes agrupados por movimiento
    """
    try:
        query = """
            SELECT 
                m.id as movimiento_id,
                m.fecha,
                m.monto_total,
                d.id as detalle_id,
                d.producto_id,
                d.descripcion,
                d.cantidad,
                d.precio_unitario,
                d.subtotal,
                d.porcentaje_iva,
                d.importe_iva,
                p.codigo as producto_codigo,
                COALESCE(p.precio, d.precio_unitario) as precio_actual
            FROM cta_cte_movimiento m
            INNER JOIN cta_cte_detalle d ON m.id = d.movimiento_id
            LEFT JOIN producto p ON d.producto_id = p.id
            WHERE m.cliente_id = :cliente_id 
            AND m.estado = 'pendiente'
            AND m.tipo = 'venta_fiada'
            ORDER BY m.fecha DESC, d.id ASC
        """
        
        result = ejecutar_query(db, query, {'cliente_id': cliente_id})
        rows = result.fetchall()
        
        # Agrupar por movimiento
        movimientos = {}
        for row in rows:
            mov_id = row.movimiento_id
            
            if mov_id not in movimientos:
                movimientos[mov_id] = {
                    'movimiento_id': mov_id,
                    'fecha': row.fecha.strftime('%d/%m/%Y %H:%M') if row.fecha else '',
                    'monto_total': float(row.monto_total),
                    'productos': []
                }
            
            movimientos[mov_id]['productos'].append({
                'detalle_id': row.detalle_id,
                'producto_id': row.producto_id,
                'producto_codigo': row.producto_codigo,
                'descripcion': row.descripcion,
                'cantidad': float(row.cantidad),
                'precio_unitario': float(row.precio_unitario),
                'subtotal': float(row.subtotal),
                'porcentaje_iva': float(row.porcentaje_iva),
                'importe_iva': float(row.importe_iva),
                'precio_actual': float(row.precio_actual) if row.precio_actual else float(row.precio_unitario),
                'subtotal_actual': float(row.cantidad) * (float(row.precio_actual) if row.precio_actual else float(row.precio_unitario))
            })
        
        return list(movimientos.values())
        
    except Exception as e:
        print(f"Error al obtener productos pendientes: {str(e)}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES PRINCIPALES - PROCESAR PAGO
# ═══════════════════════════════════════════════════════════════════════════════

def marcar_productos_como_pagados(db, detalle_ids, factura_id):
    """
    Marca productos específicos como pagados y vincula con la factura
    
    Args:
        db: Instancia de SQLAlchemy
        detalle_ids: Lista de IDs de cta_cte_detalle que se están pagando
        factura_id: ID de la factura generada
    
    Returns:
        dict con {success: bool, mensaje: str}
    """
    try:
        if not detalle_ids:
            return {'success': False, 'mensaje': 'No se especificaron productos a pagar'}
        
        # Obtener los movimientos afectados
        query_movimientos = """
            SELECT DISTINCT movimiento_id 
            FROM cta_cte_detalle 
            WHERE id IN :detalle_ids
        """
        
        # Convertir lista a formato SQL
        detalle_ids_str = '(' + ','.join(map(str, detalle_ids)) + ')'
        
        result = db.session.execute(
            text(f"SELECT DISTINCT movimiento_id FROM cta_cte_detalle WHERE id IN {detalle_ids_str}")
        )
        movimientos_ids = [row.movimiento_id for row in result]
        
        # Para cada movimiento, verificar si todos sus productos fueron pagados
        for mov_id in movimientos_ids:
            # Contar productos totales del movimiento
            query_total = """
                SELECT COUNT(*) as total 
                FROM cta_cte_detalle 
                WHERE movimiento_id = :mov_id
            """
            result_total = ejecutar_query(db, query_total, {'mov_id': mov_id})
            total_productos = result_total.fetchone().total
            
            # Contar cuántos de esos productos están en detalle_ids
            query_pagados = f"""
                SELECT COUNT(*) as pagados 
                FROM cta_cte_detalle 
                WHERE movimiento_id = :mov_id 
                AND id IN {detalle_ids_str}
            """
            result_pagados = db.session.execute(text(query_pagados), {'mov_id': mov_id})
            productos_pagados = result_pagados.fetchone().pagados
            
            # Si se pagaron todos los productos, marcar movimiento como pagado
            if total_productos == productos_pagados:
                query_update_mov = """
                    UPDATE cta_cte_movimiento 
                    SET estado = 'pagado', factura_id = :factura_id 
                    WHERE id = :mov_id
                """
                ejecutar_query(db, query_update_mov, {
                    'factura_id': factura_id,
                    'mov_id': mov_id
                }, commit=True)
        
        return {
            'success': True,
            'mensaje': 'Productos marcados como pagados correctamente'
        }
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'mensaje': f'Error al marcar productos como pagados: {str(e)}'
        }


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CONSULTA
# ═══════════════════════════════════════════════════════════════════════════════

def obtener_saldo_cliente(db, cliente_id):
    """
    Obtiene el saldo pendiente de un cliente
    
    Args:
        db: Instancia de SQLAlchemy
        cliente_id: ID del cliente
    
    Returns:
        float con el saldo pendiente
    """
    try:
        query = """
            SELECT COALESCE(SUM(monto_total), 0) as saldo
            FROM cta_cte_movimiento
            WHERE cliente_id = :cliente_id
            AND estado = 'pendiente'
            AND tipo = 'venta_fiada'
        """
        
        result = ejecutar_query(db, query, {'cliente_id': cliente_id})
        row = result.fetchone()
        
        return float(row.saldo) if row else 0.0
        
    except Exception as e:
        print(f"Error al obtener saldo: {str(e)}")
        return 0.0


def obtener_historial_cta_cte(db, cliente_id, limit=50):
    """
    Obtiene el historial completo de movimientos de un cliente
    
    Args:
        db: Instancia de SQLAlchemy
        cliente_id: ID del cliente
        limit: Cantidad máxima de registros
    
    Returns:
        Lista de movimientos
    """
    try:
        query = """
            SELECT 
                m.id,
                m.fecha,
                m.tipo,
                m.estado,
                m.monto_total,
                m.factura_id,
                f.numero as factura_numero,
                u.nombre as usuario_nombre
            FROM cta_cte_movimiento m
            LEFT JOIN factura f ON m.factura_id = f.id
            LEFT JOIN usuario u ON m.usuario_id = u.id
            WHERE m.cliente_id = :cliente_id
            ORDER BY m.fecha DESC
            LIMIT :limit
        """
        
        result = ejecutar_query(db, query, {
            'cliente_id': cliente_id,
            'limit': limit
        })
        
        movimientos = []
        for row in result:
            movimientos.append({
                'id': row.id,
                'fecha': row.fecha.strftime('%d/%m/%Y %H:%M') if row.fecha else '',
                'tipo': row.tipo,
                'estado': row.estado,
                'monto_total': float(row.monto_total),
                'factura_id': row.factura_id,
                'factura_numero': row.factura_numero,
                'usuario_nombre': row.usuario_nombre
            })
        
        return movimientos
        
    except Exception as e:
        print(f"Error al obtener historial: {str(e)}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# RUTAS API (ENDPOINTS)
# ═══════════════════════════════════════════════════════════════════════════════

@cta_cte_bp.route('/api/cta_cte/productos_pendientes/<int:cliente_id>', methods=['GET'])
def api_productos_pendientes(cliente_id):
    """
    Endpoint para obtener productos pendientes de un cliente
    GET /api/cta_cte/productos_pendientes/123
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    productos = obtener_productos_pendientes(db, cliente_id)
    saldo = obtener_saldo_cliente(db, cliente_id)
    
    return jsonify({
        'success': True,
        'productos': productos,
        'saldo_total': saldo
    })


@cta_cte_bp.route('/api/cta_cte/saldo/<int:cliente_id>', methods=['GET'])
def api_saldo_cliente(cliente_id):
    """
    Endpoint para obtener solo el saldo de un cliente
    GET /api/cta_cte/saldo/123
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    saldo = obtener_saldo_cliente(db, cliente_id)
    
    return jsonify({
        'success': True,
        'saldo': saldo
    })


@cta_cte_bp.route('/api/cta_cte/historial/<int:cliente_id>', methods=['GET'])
def api_historial_cliente(cliente_id):
    """
    Endpoint para obtener historial de movimientos
    GET /api/cta_cte/historial/123
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    historial = obtener_historial_cta_cte(db, cliente_id)
    
    return jsonify({
        'success': True,
        'historial': historial
    })


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN DE INICIALIZACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def init_cta_cte(app):
    """
    Inicializa el módulo de cuenta corriente en la app Flask
    
    Uso en app.py:
        from cta_cte import init_cta_cte
        init_cta_cte(app)
    """
    app.register_blueprint(cta_cte_bp)
    print("✅ Módulo CTA.CTE inicializado correctamente")


@cta_cte_bp.route('/cuentas_corrientes')
def vista_cuentas_corrientes():
    """Renderiza la vista principal de cuentas corrientes"""
    return render_template('ctas_ctes.html')


@cta_cte_bp.route('/api/cta_cte/resumen_general', methods=['GET'])
def api_resumen_general():
    """
    Obtiene resumen general de todas las cuentas corrientes
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    try:
        # Resumen general
        query_resumen = """
            SELECT 
                COUNT(DISTINCT CASE WHEN m.estado = 'pendiente' THEN m.cliente_id END) as clientes_con_deuda,
                COUNT(DISTINCT m.cliente_id) as clientes_activos,
                COUNT(*) as total_movimientos,
                SUM(CASE WHEN m.estado = 'pendiente' THEN m.monto_total ELSE 0 END) as total_pendiente,
                SUM(CASE WHEN m.estado = 'pagado' THEN m.monto_total ELSE 0 END) as total_cobrado,
                COUNT(CASE WHEN m.estado = 'pagado' THEN 1 END) as cantidad_pagos
            FROM cta_cte_movimiento m
        """
        
        result = ejecutar_query(db, query_resumen)
        row_resumen = result.fetchone()
        
        resumen = {
            'clientes_con_deuda': row_resumen.clientes_con_deuda or 0,
            'clientes_activos': row_resumen.clientes_activos or 0,
            'total_movimientos': row_resumen.total_movimientos or 0,
            'total_pendiente': float(row_resumen.total_pendiente or 0),
            'total_cobrado': float(row_resumen.total_cobrado or 0),
            'cantidad_pagos': row_resumen.cantidad_pagos or 0
        }
        
        # Detalle por cliente
        query_clientes = """
            SELECT 
                c.id,
                c.nombre,
                c.documento,
                COUNT(CASE WHEN m.estado = 'pendiente' THEN 1 END) as movimientos_pendientes,
                COALESCE(SUM(CASE WHEN m.estado = 'pendiente' THEN m.monto_total END), 0) as saldo_pendiente,
                MAX(m.fecha) as ultima_operacion
            FROM cliente c
            INNER JOIN cta_cte_movimiento m ON c.id = m.cliente_id
            GROUP BY c.id, c.nombre, c.documento
            ORDER BY saldo_pendiente DESC
        """
        
        result_clientes = ejecutar_query(db, query_clientes)
        
        clientes = []
        for row in result_clientes:
            clientes.append({
                'id': row.id,
                'nombre': row.nombre,
                'documento': row.documento,
                'movimientos_pendientes': row.movimientos_pendientes,
                'saldo_pendiente': float(row.saldo_pendiente),
                'ultima_operacion': row.ultima_operacion.strftime('%d/%m/%Y %H:%M') if row.ultima_operacion else None
            })
        
        return jsonify({
            'success': True,
            'resumen': resumen,
            'clientes': clientes
        })
        
    except Exception as e:
        print(f"Error en resumen general: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cta_cte_bp.route('/api/cta_cte/detalle_cliente/<int:cliente_id>', methods=['GET'])
def api_detalle_cliente(cliente_id):
    """
    Obtiene detalle completo de un cliente con todos sus movimientos
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    try:
        # Datos del cliente
        query_cliente = """
            SELECT 
                c.id,
                c.nombre,
                c.documento,
                COALESCE(SUM(CASE WHEN m.estado = 'pendiente' THEN m.monto_total END), 0) as saldo_pendiente
            FROM cliente c
            LEFT JOIN cta_cte_movimiento m ON c.id = m.cliente_id
            WHERE c.id = :cliente_id
            GROUP BY c.id, c.nombre, c.documento
        """
        
        result = ejecutar_query(db, query_cliente, {'cliente_id': cliente_id})
        row = result.fetchone()
        
        if not row:
            return jsonify({
                'success': False,
                'error': 'Cliente no encontrado'
            }), 404
        
        cliente = {
            'id': row.id,
            'nombre': row.nombre,
            'documento': row.documento,
            'saldo_pendiente': float(row.saldo_pendiente)
        }
        
        # Movimientos del cliente
        query_movimientos = """
            SELECT 
                m.id,
                m.fecha,
                m.tipo,
                m.estado,
                m.monto_total,
                m.factura_id,
                f.numero as factura_numero
            FROM cta_cte_movimiento m
            LEFT JOIN factura f ON m.factura_id = f.id
            WHERE m.cliente_id = :cliente_id
            ORDER BY m.fecha DESC
        """
        
        result_mov = ejecutar_query(db, query_movimientos, {'cliente_id': cliente_id})
        
        movimientos = []
        for row in result_mov:
            movimientos.append({
                'id': row.id,
                'fecha': row.fecha.strftime('%d/%m/%Y %H:%M') if row.fecha else '',
                'tipo': row.tipo,
                'estado': row.estado,
                'monto_total': float(row.monto_total),
                'factura_id': row.factura_id,
                'factura_numero': row.factura_numero
            })
        
        return jsonify({
            'success': True,
            'cliente': cliente,
            'movimientos': movimientos
        })
        
    except Exception as e:
        print(f"Error en detalle cliente: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cta_cte_bp.route('/api/cta_cte/productos_movimiento/<int:movimiento_id>', methods=['GET'])
def api_productos_movimiento(movimiento_id):
    """
    Obtiene los productos de un movimiento específico
    """
    from flask import current_app
    db = current_app.extensions['sqlalchemy'].db
    
    try:
        query = """
            SELECT 
                d.id,
                d.producto_id,
                d.descripcion,
                d.cantidad,
                d.precio_unitario,
                d.subtotal,
                d.porcentaje_iva
            FROM cta_cte_detalle d
            WHERE d.movimiento_id = :movimiento_id
            ORDER BY d.id
        """
        
        result = ejecutar_query(db, query, {'movimiento_id': movimiento_id})
        
        productos = []
        for row in result:
            productos.append({
                'id': row.id,
                'producto_id': row.producto_id,
                'descripcion': row.descripcion,
                'cantidad': float(row.cantidad),
                'precio_unitario': float(row.precio_unitario),
                'subtotal': float(row.subtotal),
                'porcentaje_iva': float(row.porcentaje_iva)
            })
        
        return jsonify({
            'success': True,
            'productos': productos
        })
        
    except Exception as e:
        print(f"Error al obtener productos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================================
# REEMPLAZAR DESDE LA LÍNEA 672 HASTA LA 734 EN cta_cte.py CON ESTE CÓDIGO
# ============================================================================

@cta_cte_bp.route('/api/cta_cte/exportar_excel', methods=['GET'])
def api_exportar_excel():
    """
    Exporta reporte de cuentas corrientes a Excel REAL (.xlsx)
    Mismo estilo que el reporte de ventas usando openpyxl
    """
    from flask import current_app, send_file
    import io
    from datetime import datetime
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no está instalado'}), 500
    
    db = current_app.extensions['sqlalchemy'].db
    
    try:
        print("🔍 Iniciando exportación Excel cuentas corrientes")
        
        query = """
            SELECT 
                c.id,
                c.nombre,
                c.documento,
                c.tipo_documento,
                c.telefono,
                c.email,
                COUNT(DISTINCT m.id) as movimientos_pendientes,
                COALESCE(SUM(CASE WHEN m.tipo = 'venta_fiada' THEN m.monto_total ELSE -m.monto_total END), 0) as saldo_pendiente,
                MAX(m.fecha) as ultima_operacion
            FROM cliente c
            LEFT JOIN cta_cte_movimiento m ON c.id = m.cliente_id AND m.estado = 'pendiente'
            GROUP BY c.id, c.nombre, c.documento, c.tipo_documento, c.telefono, c.email
            HAVING saldo_pendiente > 0 OR movimientos_pendientes > 0
            ORDER BY saldo_pendiente DESC
        """
        
        result = ejecutar_query(db, query)
        
        # Procesar datos
        total_adeudado = 0
        clientes_con_deuda = 0
        total_movimientos = 0
        total_clientes = 0
        
        datos_clientes = []
        for row in result:
            saldo = float(row.saldo_pendiente)
            movimientos = int(row.movimientos_pendientes)
            
            ultima_op_str = 'Sin operaciones'
            if row.ultima_operacion:
                try:
                    ultima_op_str = row.ultima_operacion.strftime('%d/%m/%Y')
                except:
                    ultima_op_str = str(row.ultima_operacion)[:10]
            
            estado = 'DEBE' if saldo > 0 else 'AL DÍA'
            
            datos_clientes.append({
                'id': row.id,
                'nombre': row.nombre,
                'tipo_documento': row.tipo_documento or 'N/A',
                'documento': str(row.documento or 'S/D'),
                'telefono': str(row.telefono or 'N/A'),
                'email': row.email or 'N/A',
                'movimientos': movimientos,
                'saldo': saldo,
                'estado': estado,
                'ultima_operacion': ultima_op_str
            })
            
            total_adeudado += saldo
            if saldo > 0:
                clientes_con_deuda += 1
            total_movimientos += movimientos
            total_clientes += 1
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas Corrientes"
        
        # Estilos (igual al reporte de ventas)
        titulo_font = Font(bold=True, size=16, color="FFFFFF")
        titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        encabezado_font = Font(bold=True, size=11, color="FFFFFF")
        encabezado_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:J1')
        celda_titulo = ws['A1']
        celda_titulo.value = 'Reporte de Cuentas Corrientes'
        celda_titulo.font = titulo_font
        celda_titulo.fill = titulo_fill
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Resumen
        fila = 4
        ws.merge_cells(f'A{fila}:J{fila}')
        ws[f'A{fila}'] = 'RESUMEN DEL PERÍODO'
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        
        fila += 1
        ws[f'A{fila}'] = 'Clientes con Movimientos:'
        ws[f'B{fila}'] = total_clientes
        ws[f'D{fila}'] = 'Total Adeudado:'
        ws[f'E{fila}'] = total_adeudado
        ws[f'E{fila}'].number_format = '$#,##0.00'
        
        fila += 1
        ws[f'A{fila}'] = 'Clientes con Deuda:'
        ws[f'B{fila}'] = clientes_con_deuda
        ws[f'D{fila}'] = 'Movimientos Pendientes:'
        ws[f'E{fila}'] = total_movimientos
        
        # Encabezados de tabla
        fila += 2
        encabezados = [
            'ID', 'Cliente', 'Tipo Doc.', 'Documento', 'Teléfono',
            'Email', 'Mov. Pendientes', 'Saldo Pendiente', 'Estado', 'Última Operación'
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=fila, column=col, value=encabezado)
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = border
        
        # Datos
        fila += 1
        for cliente in datos_clientes:
            ws.cell(row=fila, column=1, value=cliente['id'])
            ws.cell(row=fila, column=2, value=cliente['nombre'])
            ws.cell(row=fila, column=3, value=cliente['tipo_documento'])
            ws.cell(row=fila, column=4, value=cliente['documento'])
            ws.cell(row=fila, column=5, value=cliente['telefono'])
            ws.cell(row=fila, column=6, value=cliente['email'])
            ws.cell(row=fila, column=7, value=f"{cliente['movimientos']} movimientos")
            
            celda_saldo = ws.cell(row=fila, column=8, value=cliente['saldo'])
            celda_saldo.number_format = '$#,##0.00'
            
            ws.cell(row=fila, column=9, value=cliente['estado'])
            ws.cell(row=fila, column=10, value=cliente['ultima_operacion'])
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 18
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'Cuentas_Corrientes_{fecha_actual}.xlsx'
        
        print(f"✅ Excel generado exitosamente: {nombre_archivo}")
        print(f"   Total clientes: {total_clientes}")
        print(f"   Total adeudado: ${total_adeudado:,.2f}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        print(f"❌ Error generando Excel de cuentas corrientes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'connection',
            'detail': str(e)
        }), 500